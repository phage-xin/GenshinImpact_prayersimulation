import random
import openpyxl
from openpyxl.styles import Font, PatternFill
import datetime
import os


WD = os.getcwd()
''' WD = '/home/vxin/桌面/python/games/yuanshen' '''
os.chdir(WD)

ResidentAwardPoolPath = './奖池.xlsx'
WinningProbabilityPath = './星级及概率.xlsx'
LotteryRecordPath = './记录.xlsx'

TitleBlockFont = Font(bold=True)
FiveStarFont = Font(bold=True, color='00FFFF00')
FourStarFont = Font(bold=True, color='00FF00FF')
AllFill = PatternFill(fill_type='solid', fgColor='00C0C0C0')


def outprint(star, type, name):
    """打印本次抽中的奖品

    Args:
        star (int): 星级
        type (str): 类型
        name (str): 名称
    """
    if star == 5:
        color = 33
    elif star == 4:
        color = 35
    elif star == 3:
        color = 34
    else:
        color = 31
    #  print('\033[0;{};40m星级:{} 种类:{} 名称:{}\033[0m'.format(color, star, type, name))
    print('星级:{} 种类:{} 名称:{}'.format(star, type, name))

def query(project):
    """打印该抽奖项目的抽奖记录

    Args:
        project (str): 抽奖项目（常驻祈愿、角色活动祈愿、武器活动祈愿）
    """    
    if os.path.exists(LotteryRecordPath):
        wb = openpyxl.load_workbook(LotteryRecordPath)
        ws = wb[project]
        for i in range(1, len([v for v in ws.values])):
            if ws['D%d'%(i+1)].value == 5:
                color = 33
            elif ws['D%d'%(i+1)].value == 4:
                color = 35
            elif ws['D%d'%(i+1)].value == 3:
                color = 34
            else:
                color = 31
            print('\033[0;{};40m第{}次,抽到{}星{},名称:{}\033[0m'.format(
                  color, ws['E%d'%(i+1)].value, ws['D%d'%(i+1)].value,
                         ws['C%d'%(i+1)].value, ws['B%d'%(i+1)].value))
    else:
        print('未找到 %s 文件!'%(LotteryRecordPath))

def get_frequency(project):
    """查询抽各星级对应的概率（%）、保底次数、为本期物品概率（%）。加载当前工作目录下的文件 星级及概率.xlsx

    Args:
        project (str): 抽奖项目（常驻祈愿、角色活动祈愿、武器活动祈愿）

    Returns:
        dict: {星级:list(概率（%）, 保底, 为本期物品概率（%）)}
    """    
    wb = openpyxl.load_workbook(WinningProbabilityPath)
    ws = wb[project]
    d = {}
    for y in (2, 3, 4):
        for x in ('A', 'B', 'C', 'D'):
            if x == 'A':
                star = ws['%s%d'%(x, y)].value
            elif x == 'B':
                frequency = ws['%s%d'%(x, y)].value
            elif x == 'C':
                minimum_times = ws['%s%d'%(x, y)].value
            else:
                ipota = ws['%s%d'%(x, y)].value  # Item probability of this activity
        d.update({star:[frequency, minimum_times, ipota]})
    return d

def get_jackpot(project):
    """导入奖池内容。加载当前工作目录下的文件 奖池.xlsx

    Args:
        project (str): 抽奖项目（常驻祈愿、角色活动祈愿、武器活动祈愿）

    Returns:
        tuple: 三个列表[星级， 类型， 名称]
    """    
    wb = openpyxl.load_workbook(ResidentAwardPoolPath)
    ws = wb[project]
    five_star = []
    four_stat = []
    three_star = []
    t = 0
    for row in ws.values:
        t += 1
        if t == 1:
            pass
        else:
            tl = tuple(row)
            if tl[0] == 5:
                five_star.append(tl)
            elif tl[0] == 4:
                four_stat.append(tl)
            else:
                three_star.append(tl)
    return (five_star, four_stat, three_star)

def keep_records(project, star, type, name, remakes):
    """保存抽奖记录。保存文当前工作目录下的 记录.xlsx

    Args:
        project (str): 抽奖项目（常驻祈愿、角色活动祈愿、武器活动祈愿）
        star (int): 星级
        type (str): 类型
        name (str): 名称
        remakes (str)): 限定角色标记奖池备注
    """    
    if os.path.exists(LotteryRecordPath):
        wb = openpyxl.load_workbook(LotteryRecordPath)
        ws = wb[project]
        p = len([v for v in ws.values])  # p为列表的行数
        ws['A%d'%(p+1)] = datetime.datetime.today()  # A列为时间
        ws['A%d'%(p+1)].number_format = "yyyy-mm-dd hh:mm:ss"
        ws['B%d'%(p+1)] = name  # B列为名称
        ws['C%d'%(p+1)] = type  # C列为种类
        ws['D%d'%(p+1)] = int(star)  # D列为星级
        if p == 1:
            ws['E%d'%(p+1)] = 1  # E列为总次数
            ws['F%d'%(p+1)] = 1  # F列为距离上次5星的次数
            ws['G%d'%(p+1)] = 1  # G列为距离上次4星的次数
        else:
            ws['E%d'%(p+1)] = ws['E%d'%(p)].value + 1
            if ws['D%d'%(p)].value == 5:
                ws['F%d'%(p+1)] = 1
                ws['G%d'%(p+1)] = 1
            elif ws['D%d'%(p)].value == 4:
                ws['F%d'%(p+1)] = ws['F%d'%(p)].value + 1
                ws['G%d'%(p+1)] = 1
            else:
                ws['F%d'%(p+1)] = ws['F%d'%(p)].value + 1
                ws['G%d'%(p+1)] = ws['G%d'%(p)].value + 1
        ws['H%d'%(p+1)] = remakes  # H列为备注
        for i in range(ord('A'), ord('I')):
            ws['%s%d'%(chr(i), p+1)].fill = AllFill
            if star == 5:
                ws['%s%d'%(chr(i), p+1)].font = FiveStarFont
            elif star == 4:
                ws['%s%d'%(chr(i), p+1)].font = FourStarFont
            else:
                pass
        wb.save(LotteryRecordPath)
    else:
        wb = openpyxl.Workbook()
        ws1 = wb.create_sheet("常驻祈愿")
        ws2 = wb.create_sheet("角色活动祈愿")
        ws3 = wb.create_sheet("武器活动祈愿")
        del wb['Sheet']
        for i in (1, 2, 3):
            exec('ws%d.column_dimensions["A"].width = 20'%(i))
            exec('ws%d.column_dimensions["B"].width = 15'%(i))
            exec('ws%d.column_dimensions["F"].width = 18'%(i))
            exec('ws%d.column_dimensions["G"].width = 18'%(i))
            exec('ws%d["A1"] = "时间"'%(i))
            exec('ws%d["B1"] = "名称"'%(i))
            exec('ws%d["C1"] = "类别"'%(i))
            exec('ws%d["D1"] = "星级"'%(i))
            exec('ws%d["E1"] = "总次数"'%(i))
            exec('ws%d["F1"] = "距离上次5星的次数"'%(i))
            exec('ws%d["G1"] = "距离上次4星的次数"'%(i))
            exec('ws%d["H1"] = "备注"'%(i))
            for j in range(ord('A'), ord('I')):
                exec('ws%d["%s1"].font = TitleBlockFont'%(i, chr(j)))
                exec('ws%d["%s1"].fill = AllFill'%(i, chr(j)))
            exec('ws%d.freeze_panes = "A2"'%(i))
        wb.save(LotteryRecordPath)
        keep_records(project, star, type, name, remakes)

def minimum_mark(project, x=0):
    """返回5星及4星的保底标志

    Args:
        project (str): 抽奖项目（常驻祈愿、角色活动祈愿、武器活动祈愿）
        x (int): 5星保底次数

    Returns:
        tuple: 5星的标记， 4星的标记[, 上次5星角色名称， 上次4星角色名称]
    """    
    if os.path.exists(LotteryRecordPath):
        wb = openpyxl.load_workbook(LotteryRecordPath)
        ws = wb[project]
        p = len([v for v in ws.values])
        five = ws['F%d'%(p)].value
        four = ws['G%d'%(p)].value
        if project == '常驻祈愿':
            return (five, four)
        else:
            def last(x):
                lfive = ['N']
                lfour = ['N']
                if p <= x:
                    for row in ws.values:
                        l = list(row)
                        if l[3] == 5:
                            lfive.append(l[1])
                        elif l[3] == 4:
                            lfour.append(l[1])
                else:
                    for row in list(ws.values)[p-90:]:
                        l = list(row)
                        if l[3] == 5:
                            lfive.append(l[1])
                        elif l[3] == 4:
                            lfour.append(l[1])
                return (lfive[-1], lfour[-1])
            if project == '角色活动祈愿':
                (last_five, last_four) = last(x)
                return (five, four, last_five, last_four)
            elif project == '武器活动祈愿':
                wb1 = openpyxl.load_workbook(ResidentAwardPoolPath)
                ws1 = wb['武器活动祈愿']
                m = ws1['G1'].value
                (last_five, last_four) = last(x)
                return (five, four, last_five, last_four, m)
    else:
        if project == '常驻祈愿':
            return (0, 0)
        elif project == '角色活动祈愿':
            return (0, 0, 0, 0)
        elif project == '武器活动祈愿':
            wb1 = openpyxl.load_workbook(ResidentAwardPoolPath)
            ws1 = wb1['武器活动祈愿']
            m = ws1['G1'].value
            return (0, 0, 0, 0, m)
        else:
            pass

def get_jackpot2(project, remakes):
    (five_star1, four_star1, three_star1) = get_jackpot('常驻祈愿')
    if project == '常驻祈愿':
        return (five_star1, four_star1, three_star1)
    else:
        if project == '角色活动祈愿':
            for f in get_jackpot(project)[0]:
                if f[2] == remakes[:-1]:
                    five_star2 = [f]
            four_star2 = get_jackpot(project)[1]
            for a in five_star1:
                if a[0] == 5 and a[1] == '武器':
                    five_star1.remove(a)
            for a in four_star2:
                if a in four_star1:
                    four_star1.remove(a)
        elif project == '武器活动祈愿':
            five_star2 = []
            for f in get_jackpot(project)[0]:
                five_star2.append(f)
            four_star2 = get_jackpot(project)[1]
            for a in five_star1:
                if a[0] == 5 and a[1] == '角色':
                    five_star1.remove(a)
            for a in four_star2:
                if a in four_star1:
                    four_star1.remove(a)
        return (five_star1, four_star1, three_star1, five_star2, four_star2)

def ger_frequency2(project):
    d = get_frequency(project)  # star:[frequency, minimum_times, ipota]
    three_frequency = int(d[3][0]*10)
    (five_frequency, five_minimun_times) = (int(d[5][0]*10), int(d[5][1]))
    (four_frequnecy, four_minimun_times) = (int(d[4][0]*10), int(d[4][1]))
    if project == '常驻祈愿':
        return (five_frequency, five_minimun_times, four_frequnecy, four_minimun_times, three_frequency)
    else:
        (five_frequency, five_minimun_times, five_ipota) = (int(d[5][0]*10), int(d[5][1]), int(d[5][2]*10))
        (four_frequnecy, four_minimun_times, four_ipota) = (int(d[4][0]*10), int(d[4][1]), int(d[5][2]*10))
        return (five_frequency, five_minimun_times, five_ipota, four_frequnecy, four_minimun_times, four_ipota, three_frequency, d[5][2])

def luck_draw(project, jackpot2, frequency2):
    """抽奖模块

    Args:
        project (str): 抽奖项目（常驻祈愿、角色活动祈愿、武器活动祈愿）
        jackpot2 (tupe): get_jackpot2的返回值
        frequency2 (tupe): get_frequency2的返回值

    Returns:
        list: 星级， 类别， 名称
    """    
    if project == '常驻祈愿':
        (five_star1, four_star1, three_star1) = jackpot2
        (five_mark, four_mark) = minimum_mark(project)
        (five_frequency, five_minimun_times, four_frequnecy, four_minimun_times, three_frequency) = frequency2
        if five_mark == five_minimun_times - 1:  # 五星保底
            return random.choice(five_star1)
        else:
            if four_mark == four_minimun_times - 1:  # 四星保底
                return random.choice(four_star1)
            else:  # 常规
                x = random.randint(1, five_frequency+four_frequnecy+three_frequency)
                if x in range(1, five_frequency+1):
                    return random.choice(five_star1)
                elif x in range(five_frequency+1, five_frequency+four_frequnecy+1):
                    return random.choice(four_star1)
                elif x in range(five_frequency+four_frequnecy+1, five_frequency+four_frequnecy+three_frequency+1):
                    return random.choice(three_star1)
                else:
                    return ['!', '!', '!']
    else:
        (five_star1, four_star1, three_star1, five_star2, four_star2) = jackpot2
        (five_frequency, five_minimun_times, five_ipota, four_frequnecy, four_minimun_times, four_ipota, three_frequency, d) = frequency2
        if project == '角色活动祈愿':
            (five_mark, four_mark, last_five, last_four) = minimum_mark(project, d)
            if five_mark == five_minimun_times - 1:  # 五星保底
                if last_five in [n[2] for n in five_star1]:  # 本期五星保底
                    return random.choice(five_star2)
                else:  # 小保底
                    x = random.randint(1, 1000)
                    if x in range(1, five_ipota+1):
                        return random.choice(five_star2)
                    else:
                        return random.choice(five_star1)
            else:
                if four_mark == four_minimun_times - 1:  # 四星保底
                    if last_four in [n[2] for n in four_star1]:  # 本期四星保底
                        return random.choice(four_star2)
                    else:  # 常规四星保底
                        x = random.randint(1, 1000)
                        if x in range(1, four_ipota+1):
                            return random.choice(four_star2)
                        else:
                            return random.choice(four_star1)
                else:  # 常规
                    x = random.randint(1, five_frequency+four_frequnecy+three_frequency)
                    if x in range(1, five_frequency+1):
                        i = random.randint(1, 1000)
                        if x in range(1, five_ipota+1):
                            return random.choice(five_star2)
                        else:
                            return random.choice(five_star1)
                    elif x in range(five_frequency+1, four_frequnecy+five_frequency+1):
                        i = random.randint(1, 1000)
                        if i in range(1, four_ipota):
                            return random.choice(four_star2)
                        else:
                            return random.choice(four_star1)
                    elif x in range(four_frequnecy+five_frequency+1, three_frequency+four_frequnecy+five_frequency+1):
                        return random.choice(three_star1)
                    else:
                        return['!', '!', '!']
        elif project == '武器活动祈愿':
            (five_mark, four_mark, last_five, last_four, m_value) = minimum_mark(project, d)
            wb = openpyxl.load_workbook(ResidentAwardPoolPath)
            ws = wb[project]
            od_weapon = ws['E1'].value
            if five_mark == five_minimun_times - 1:  # 五星保底
                if m_value == 2:  # 定轨保底
                    revise_weapon(2)
                    return [5, '武器', od_weapon]
                else:  # 非定轨保底
                    if last_five in [n[2] for n in five_star1]:  # 本期五星保底
                        j = random.choice(five_star2)
                        if j[2] == od_weapon:
                            revise_weapon(2)
                        return j
                    else:  # 常规五星保底
                        x = random.randint(1, 1000)
                        if x in range(1, five_ipota+1):
                            j = random.choice(five_star2)
                            if j[2] == od_weapon:
                                revise_weapon(2)
                            else:
                                revise_weapon(1)
                            return j
                        else:
                            revise_weapon(1)
                            return random.choice(five_star1)
            else:
                if four_mark == four_minimun_times - 1:  # 四星保底
                    if last_four in [n[2] for n in four_star1]:  # 本期四星保底
                        return random.choice(four_star2)
                    else:  # 常规四星保底
                        x = random.randint(1, 1000)
                        if x in range(1, four_ipota+1):
                            return random.choice(four_star2)
                        else:
                            return random.choice(four_star1)
                else:  # 常规
                    x = random.randint(1, five_frequency+four_frequnecy+three_frequency)
                    if x in range(1, five_frequency+1):
                        i = random.randint(1, 1000)
                        if x in range(1, five_ipota+1):
                            j = random.choice(five_star2)
                            if j[2] == od_weapon:
                                revise_weapon(2)
                            return j
                        else:
                            revise_weapon(1)
                            return random.choice(five_star1)
                    elif x in range(five_frequency+1, four_frequnecy+five_frequency+1):
                        i = random.randint(1, 1000)
                        if i in range(1, four_ipota):
                            return random.choice(four_star2)
                        else:
                            return random.choice(four_star1)
                    elif x in range(four_frequnecy+five_frequency+1, three_frequency+four_frequnecy+five_frequency+1):
                        return random.choice(three_star1)
                    else:
                        return['!', '!', '!']

def revise_weapon(mode):
    """武器定轨操作

    Args:
        mode (int): 0：取消定轨；1：正常；2.命值清零；3：修改定轨
    """    
    wb = openpyxl.load_workbook(ResidentAwardPoolPath)
    ws = wb['武器活动祈愿']
    if mode == 0:
        ws['E1'] = None
        ws['G1'] = None
    elif mode == 1:
        if ws['G1'].value != None and ws['E1'].value != None:
            ws['G1'] = ws['G1'].value + 1
        else:
            pass
    elif mode == 2:
        ws['G1'] = 0
    elif mode == 3:
        ws['G1'] = 0
        nl = []
        for row in ws.values:
            l = list(row)
            if  l[0] == 5:
                nl.append(l[2])
        def ch():
            t = 1
            for n in nl:
                print('%d : %s'%(t, n))
                t += 1
            c = int(input('请选择:'))
            if c-1 in range(len(l)):
                ws['E1'] = nl[c-1]
            else:
                ch()
        ch()
    wb.save(ResidentAwardPoolPath)

def fecondary_menu(project, remakes=None):  # 二级菜单
    secondary_cycle = 1
    while secondary_cycle:
        lis = get_jackpot2(project, remakes)
        tup = ger_frequency2(project)
        print('1.来十发\n2.来一发\n3.查询抽奖记录\n4.返回上一级\n5.退出')
        choice = int(input('请选择:'))
        if choice == 1:
            for t in range(10):
                l = luck_draw(project, lis, tup)
                (star, typ, name) = (l[0], l[1], l[2])
                keep_records(project, star, typ, name, remakes)
                outprint(star, typ, name)
        elif choice == 2:
            l = luck_draw(project, lis, tup)
            (star, typ, name) = (l[0], l[1], l[2])
            keep_records(project, star, typ, name, remakes)
            outprint(star, typ, name)
        elif choice == 3:
            query(project)
        elif choice == 4:
            secondary_cycle = 0
        elif choice == 5:
            exit()
        else:
            continue

def main():  # 主菜单
    print('!!!请将 "%s, %s" 文件移动在:%s!!!\n\n'%(
        ResidentAwardPoolPath[2::], WinningProbabilityPath[2::], WD))
    print('='*20 + '欢迎使用原神抽奖模拟系统' + '='*20)
    primary_cycle = 1
    while primary_cycle:
        print('1.常驻祈愿\n2.角色活动祈愿\n3.武器活动祈愿\n4.退出')
        a = int(input('请选择:'))
        if a in (1, 3):
            if a == 1:
                Project = '常驻祈愿'
            else:
                Project = '武器活动祈愿'
                while 1:
                    wb = openpyxl.load_workbook(ResidentAwardPoolPath)
                    ws = wb['武器活动祈愿']
                    od_weapon = ws['E1'].value
                    numerical_value = ws['G1'].value
                    if od_weapon == None and numerical_value == None:
                        print('当前无定轨!')
                    else:
                        print('当前定轨武器:%s\n定轨命值为:%d'%(od_weapon, numerical_value))
                    print('1.继续\n2.修改定轨\n3.取消定轨')
                    c = int(input('请选择:'))
                    if c == 1:
                        break
                    elif c == 2:
                        revise_weapon(3)
                        continue
                    elif c == 3:
                        revise_weapon(0)
                        continue
                    else:
                        continue
            fecondary_menu(Project)
        elif a == 2:
            Project = '角色活动祈愿'
            wb = openpyxl.load_workbook(ResidentAwardPoolPath)
            ws = wb[Project]
            nl = []
            for row in ws.values:
                l = list(row)
                if  l[0] == 5:
                    nl.append(l[2])
            def choice():
                t = 1
                for n in nl:
                    print('%d : %s池'%(t, n))
                    t += 1
                c = int(input('请选择:'))
                if c-1 in range(len(l)):
                    fecondary_menu(Project, remakes='%s'%(nl[c-1]+'池'))
                else:
                    choice()
            choice()
        elif a == 4:
            primary_cycle = 0
        else:
            continue


if __name__ == '__main__':
    main()
